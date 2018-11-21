import openpyxl

def read_row_values(row):
    row_values=[]
    for cell in row:
        row_values.append(cell.value)
    return row_values

def read_shipping_values():
    wb = openpyxl.load_workbook("magento2.xlsx")
    ws = wb.active

    header_row = ws[1]
    header_values = read_row_values(header_row)

    shipping_info_list=[]

    for row in ws.iter_rows(min_row=2):
        row_values=read_row_values(row)

        shipping_info={}

        for index in range(0,len(header_values)):
            shipping_info[header_values[index]] = row_values[index]
        shipping_info_list.append(shipping_info)

    return shipping_info_list


# wb = openpyxl.load_workbook("magento2.xlsx")
# ws = wb.active

# if ws["A1"].value == "username":
#     username = ws["A2"]

# if ws["B1"].value == "firstname":
#     firstname = ws["B2"]

# if ws["C1"].value == "lastname":
#     lastname = ws["C2"]

# if ws["D1"].value == "company":
#     company = ws["D2"]

# if ws["E1"].value == "street":
#     street = ws["E2"]

# if ws["F1"].value == "city":
#     city = ws["F2"]

# if ws["G1"].value == "state":
#     state = ws["G2"]

# if ws["H1"].value == "postcode":
#     postcode = ws["H2"]

# if ws["I1"].value == "country_id":
#     country_id = ws["I2"]

# if ws["J1"].value == "telephone":
#     phonenumber = ws["J2"]

