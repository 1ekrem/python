import datetime as dt
import matplotlib.pyplot as plt 
from matplotlib import style
import pandas as pd
import pandas_datareader.data as data
from pandas_datareader._utils import RemoteDataError
import os
import openpyxl
import time
import ta


def get_stock_prices(ticker):
    data_source = 'yahoo'
    start = dt.datetime(2018, 1 ,1)
    end = dt.datetime.today()
    df= data.DataReader(ticker, data_source, start, end)
    df.to_csv('C:\\PythonClass\\ta-ticker\\{}.csv'.format(ticker))

def tech_analysis(ticker):
    df = pd.read_csv('C:\\PythonClass\\ta-ticker\\{}.csv'.format(ticker), parse_dates=True, index_col=0)
    df['200ma'] = df['Adj Close'].rolling(window=200, min_periods=0).mean()
    df['100ma'] = df['Adj Close'].rolling(window=100, min_periods=0).mean()
    df['50ma'] = df['Adj Close'].rolling(window=50, min_periods=0).mean()
    df['20ma'] = df['Adj Close'].rolling(window=20, min_periods=0).mean()
    df['10ma'] = df['Adj Close'].rolling(window=10, min_periods=0).mean()
    df['5ma'] = df['Adj Close'].rolling(window=5, min_periods=0).mean()

    #RSI
    delta = df['Close'].diff()
    window = 14
    up_days = delta.copy()
    up_days[delta<=0]=0.0
    down_days = abs(delta.copy())
    down_days[delta>0]=0.0
    RS_up = up_days.rolling(window).mean()
    RS_down = down_days.rolling(window).mean()
    rsi= 100-100/(1+RS_up/RS_down)
    df['RSI']=rsi

    #CCI
    ndays=21
    TP = (df['High'] + df['Low'] + df['Close']) / 3 
    df['CCI'] = pd.Series((TP - TP.rolling(ndays).mean()) / (0.014 * TP.rolling(ndays).std()),name = 'CCI')
    
    # VOLUME INDICATORS
    df['moneyIndexFlow'] = ta.volume.money_flow_index(df['High'],df['Low'],df['Close'],df['Volume'])
    df['volumePriceTrend'] = ta.volume.volume_price_trend(df['Close'],df['Volume'])
    df['VWAP'] = ta.volume.volume_weighted_average_price(df['High'],df['Low'],df['Close'],df['Volume'])
    df['MarketPressure'] = ta.volume.force_index(df['Close'],df['Volume'])/100

    # MOMENTUM INDICATORS
    df['RSI-2'] = ta.momentum.rsi(df['Close'])
    df['AwesomeOscillator'] = ta.momentum.awesome_oscillator(df['High'],df['Low'])
    df['RateOfChageINC'] = ta.momentum.roc(df['Close'])
    
    # TREND INDICATORS
    df['BearishVortex'] = ta.trend.vortex_indicator_neg(df['High'],df['Low'],df['Close'])
    df['BullishVortex'] = ta.trend.vortex_indicator_pos(df['High'],df['Low'],df['Close'])
    df['DownTrend'] = ta.trend.psar_down_indicator(df['High'],df['Low'],df['Close'])
    df['UpTrend'] = ta.trend.psar_up_indicator(df['High'],df['Low'],df['Close'])

    #Save analysis file and remove the historic data
    df.tail(250).to_excel('C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx'.format(ticker))
    os.remove('C:\\PythonClass\\ta-ticker\\{}.csv'.format(ticker))

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow+1,1):
        rowSelected=[]
        for j in range(startCol, endCol+1,1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow=0 
    for i in range(startRow, endRow+1, 1):
        countCol=0
        for j in range(startCol, endCol+1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createAnalysis(ticker):
    print("Processing {}...".format(ticker))
    
    wb = openpyxl.load_workbook("C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx".format(ticker))
    sheet = wb["Sheet1"]

    template = openpyxl.load_workbook("C:\\PythonClass\\AlgoTrading\\AnalysisTemplate.xlsx")
    temp_sheet=template["result"]
    
    selectedRange3 = copyRange(1,2,26,251,sheet)
    pasteRange(1,2,26,251,temp_sheet,selectedRange3)

    template.save("C:\\Users\\ekrem\\Desktop\\Trade Analysis\\{}_TAanalysis.xlsx".format(ticker))
    print("Report is generated for {}! Please check your Trade Analysis folder on your desktop.\n".format(ticker))

    os.remove('C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx'.format(ticker))
 