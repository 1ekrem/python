import openpyxl
import time
import os

ticker = input(str("Enter ticker: "))
today=time.strftime("%Y%m%d-%H%M")

wb = openpyxl.load_workbook("C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx".format(ticker))
sheet = wb["Sheet1"]

template = openpyxl.load_workbook("C:\\PythonClass\\AlgoTrading\\AnalysisTemplate.xlsx")
temp_sheet=template["result"]

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow+1,1):
        rowSelected=[]
        for j in range(startCol, endCol+1,1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow=0 
    for i in range(startRow, endRow+1, 1):
        countCol=0
        for j in range(startCol, endCol+1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createAnalysis():
    print("Processing...")
    selectedRange = copyRange(1,2,13,251,sheet)
    selectedRange2 = copyRange(14,2,15,251,sheet)
    pasteRange(1,2,13,251,temp_sheet,selectedRange)
    pasteRange(32,2,33,251,temp_sheet,selectedRange2)
    template.save("C:\\Users\\ekrem\\Desktop\\Trade Analysis\\{}_TAanalysis.xlsx".format(ticker))
    print("Report is generated for {}! Please check your Desktop-Trade Analysis folder".format(ticker))
    os.remove('C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx'.format(ticker))
    print("NOTE: Moving Averages file is removed! ta-ticker folder should be empty")

createAnalysis()