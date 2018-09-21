# 1-Create a workbook
# Go to active sheet
# add a value
# Create another worksheet
# add value to new worksheet
# Save 

import openpyxl

# open a new workbook
wb = openpyxl.Workbook()

# Go to the active sell
ws= wb.active

ws['A1'] = 42
# ws.append(['D8','D9','D10'])

wb.create_sheet('Ekrem2',1)

ws_ekrem=wb['Ekrem2']

ws_ekrem['A1'] = 1989

wb.save("ekrem_info2.xlsx")