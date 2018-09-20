import openpyxl

#Open the workbook
first = openpyxl.load_workbook('first.xlsx')
first_ws = first.active

second = openpyxl.load_workbook('second.xlsx')
second_ws = second.active

result = openpyxl.load_workbook('result.xlsx')
result_ws = result.active

comparison_cells = ['B1', 'B2', 'B3']

for ref in comparison_cells:
    first_val = first_ws[ref].value
    second_val = second_ws[ref].value
    result_ws[ref]= "Pass" if first_val != second_val else "Fail"
        
        #print("Non-matchin value!", ref, first_val, second_val)

print("Comparison finished")

result.save('result.xlsx')
print("Results are stored in result sheet")