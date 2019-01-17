import openpyxl
import time

today=time.strftime("%Y%m%d-%H%M")

wb = openpyxl.load_workbook("GradeSample.xlsx")
sheet = wb["Grades"] 

template = openpyxl.load_workbook("Section12Grades.xlsx")
temp_sheet = template["Sheet1"] 

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow+1,1):
        rowSelected=[]
        for j in range(startCol, endCol+1,1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow=0
    for i in range(startRow, endRow+1, 1):
        countCol=0
        for j in range(startCol, endCol+1, 1):            
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createData():
    print("Processing...")
    selectedRange = copyRange(1,2,4,14,sheet)
    pastingRange = pasteRange(1,3,4,15,temp_sheet,selectedRange)
    template.save("Section12Grades{}.xlsx".format(today))
    print("Range copied and pasted!")


createData()