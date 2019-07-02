import openpyxl

wb = openpyxl.load_workbook("first.xlsx")
ws = wb.active

ws["B3"].value=ws["B1"].value + ws["B2"].value

print(ws["B3"].value)

