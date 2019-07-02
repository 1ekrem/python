import openpyxl
import time


today=time.strftime("%Y%m%d-%H%M")

wb = openpyxl.load_workbook("TSLA_analysis.xlsx")
sheet = wb["Sheet1"]

template = openpyxl.load_workbook("AnalysisTemplate.xlsx")
temp_sheet=template["result"]

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow+1,1):
        rowSelected=[]
        for j in range(startCol, endCol+1,1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow=0
    for i in range(startRow, endRow+1, 1):
        countCol=0
        for j in range(startCol, endCol+1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createAnalysis():
    print("Processing...")
    selectedRange = copyRange(1,2,13,251,sheet)
    pastingRange = pasteRange(1,2,13,251,temp_sheet,selectedRange)
    template.save("TA_analysis.xlsx")
    print("Range copied and pasted!")

createAnalysis()