import openpyxl

xlsx = openpyxl.Workbook()
ws = xlsx.active

for i in range (1,11):
    ws['B' + str(i)] = "1 * "+str(i)
    ws['C' + str(i)] = 1*i

xlsx.save('multiplication_result.xlsx')



