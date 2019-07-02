import openpyxl
import time

today=time.strftime("%Y%m%d-%H%M")


#File to be copied
wb = openpyxl.load_workbook("GradeSample.xlsx") #Add file name
sheet = wb["Grades"] #Add Sheet name

#File to be pasted into
template = openpyxl.load_workbook("Section12Grades.xlsx") #Add file name
temp_sheet = template["Sheet1"] #Add Sheet name

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createData():
    print("Processing...")
    selectedRange = copyRange(1,2,4,14,sheet)
    pastingRange = pasteRange(1,3,4,15,temp_sheet,selectedRange)
    template.save("Section12Grades{}.xlsx".format(today))
    print("Range copied and pasted!")


createData()