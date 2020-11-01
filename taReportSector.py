import datetime as dt
import matplotlib.pyplot as plt 
from matplotlib import style
import pandas as pd
import pandas_datareader.data as data
import os
import openpyxl
import time


def get_stock_prices(ticker):
    data_source = 'yahoo'
    start = dt.datetime(2018, 1 ,1)
    end = dt.datetime.today()
    df= data.DataReader(ticker, data_source, start, end)
    df.to_csv('C:\\PythonClass\\ta-ticker\\{}.csv'.format(ticker))
    #print("Stock prices of {} is pulled!".format(ticker))

def tech_analysis(ticker):
    df = pd.read_csv('C:\\PythonClass\\ta-ticker\\{}.csv'.format(ticker), parse_dates=True, index_col=0)
    df['200ma'] = df['Adj Close'].rolling(window=200, min_periods=0).mean()
    df['100ma'] = df['Adj Close'].rolling(window=100, min_periods=0).mean()
    df['50ma'] = df['Adj Close'].rolling(window=50, min_periods=0).mean()
    df['20ma'] = df['Adj Close'].rolling(window=20, min_periods=0).mean()
    df['10ma'] = df['Adj Close'].rolling(window=10, min_periods=0).mean()
    df['5ma'] = df['Adj Close'].rolling(window=5, min_periods=0).mean()
    #print("Moving Averages are created!")
    #RSI
    delta = df['Close'].diff()
    window = 14
    up_days = delta.copy()
    up_days[delta<=0]=0.0
    down_days = abs(delta.copy())
    down_days[delta>0]=0.0
    RS_up = up_days.rolling(window).mean()
    RS_down = down_days.rolling(window).mean()
    rsi= 100-100/(1+RS_up/RS_down)
    df['RSI']=rsi
    #print("Relative Strenght Index of 14 is created!")
    #CCI
    ndays=21
    TP = (df['High'] + df['Low'] + df['Close']) / 3 
    df['CCI'] = pd.Series((TP - TP.rolling(ndays).mean()) / (0.014 * TP.rolling(ndays).std()),name = 'CCI')
    #print("Commodity Channel Index of 21 is created!")
    #Save analysis file and remove the historic data
    df.tail(250).to_excel('C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx'.format(ticker))
    os.remove('C:\\PythonClass\\ta-ticker\\{}.csv'.format(ticker))
    #print("NOTE: Stock prices data file is removed!")

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow+1,1):
        rowSelected=[]
        for j in range(startCol, endCol+1,1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow=0 
    for i in range(startRow, endRow+1, 1):
        countCol=0
        for j in range(startCol, endCol+1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createAnalysis(ticker):
    print("Processing...")
    
    #today=time.strftime("%Y%m%d-%H%M")

    wb = openpyxl.load_workbook("C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx".format(ticker))
    sheet = wb["Sheet1"]

    template = openpyxl.load_workbook("AnalysisTemplate.xlsx")
    temp_sheet=template["result"]
    
    selectedRange = copyRange(1,2,13,251,sheet)
    selectedRange2 = copyRange(14,2,15,251,sheet)
    pasteRange(1,2,13,251,temp_sheet,selectedRange)
    pasteRange(32,2,33,251,temp_sheet,selectedRange2)
    template.save("C:\\Users\\ekrem\\Desktop\\Trade Analysis\\{}_TAanalysis.xlsx".format(ticker))
    #print("Report is generated for {}! Please check your Desktop\Trade Analysis folder".format(ticker))
    os.remove('C:\\PythonClass\\ta-ticker\\{}_SMAanalysis.xlsx'.format(ticker))
    #print("NOTE: Moving Averages file is removed! ta-ticker folder should be empty")
