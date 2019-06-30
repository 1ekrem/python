from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import difflib
from openpyxl import load_workbook

driver = webdriver.Chrome()
driver.maximize_window()

URL = "https://freecrm.com/"
driver.get(URL)
time.sleep(3)
driver.find_element_by_xpath("//span[text()='Log In']").click()
time.sleep(3)

driver.find_element_by_css_selector('input[name="email"]').send_keys("burhannyc@aol.com")
driver.find_element_by_css_selector('input[name="password"]').send_keys("Selenium123")
driver.find_element_by_css_selector('.ui.fluid.large.blue').click()
time.sleep(3)

#Hard coded variables for home page tabs
org_tabs=['Home', 'Calendar','Contacts','Companies','Deals', 'Tasks', 'Cases', 
'Calls', 'Documents', 'Email', 'Campaigns']

#Dynamic values pulled from excel
org_tabs2=[]

# set file path
filepath="C:/PythonClass/freecrm/utilities/left_tabs.xlsx"
# load the excel file 
wb=load_workbook(filepath)
# select active cells
sheet=wb.active
# iterate over all rows
for i in range(1,12):
    # iterate over all columns
    for j in range(1,2):
        # get particular cell value    
        cell_obj=sheet.cell(row=i,column=j)
    org_tabs2.append(cell_obj.value)
    
web_tabs=[]

for item in driver.find_elements_by_css_selector('.item'):
    str(item).replace(' ', '')
    web_tabs.append(item)

s = set(web_tabs)
web_tabs_set = [x for x in org_tabs if x not in s]

if org_tabs2 == web_tabs_set: 
    assert(org_tabs2 == web_tabs_set)
    print ("The webelements and values from excel are identical") 
else : 
    print ("The webelements and values from excel are not identical") 

driver.quit()
